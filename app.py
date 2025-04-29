from flask import Flask, request, jsonify, render_template, redirect, url_for, session, flash, send_file
import json
import os
import calendar
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import shutil
import tempfile
import hashlib
import secrets
import jdatetime as jdt
import locale

app = Flask(__name__, template_folder='templates', static_folder='static')
app.secret_key = secrets.token_hex(16)  # Generate a random secret key

# Set jdatetime locale to Persian
try:
    # Ensure the locale exists on the system; this might vary
    # Use 'fa_IR.UTF-8' or similar if 'fa_IR' doesn't work
    locale.setlocale(locale.LC_ALL, 'fa_IR') 
except locale.Error:
    print("Warning: Persian locale 'fa_IR' not found. Using default locale.")
    # Fallback or handle error as needed
jdt.set_locale(jdt.FA_LOCALE)
PERSIAN_MONTH_NAMES = [
    "", "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور", 
    "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"
]
PERSIAN_DAY_NAMES = [
    "شنبه", "یکشنبه", "دوشنبه", "سه‌شنبه", "چهارشنبه", "پنجشنبه", "جمعه"
]

# Configuration
DATA_DIR = 'data'
if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)

ENGINEERS_FILE = os.path.join(DATA_DIR, 'engineers.json')
SCHEDULES_FILE = os.path.join(DATA_DIR, 'schedules.json')
USERS_FILE = os.path.join(DATA_DIR, 'users.json')
WORKPLACES = ["Studio Hispan", "Studio Press", "Nodal", "Engineer Room"]
SHIFTS = ["Shift 1", "Shift 2", "Shift 3"]

# Initialize data files if they don't exist
if not os.path.exists(ENGINEERS_FILE):
    with open(ENGINEERS_FILE, 'w') as f:
        json.dump([], f)

if not os.path.exists(SCHEDULES_FILE):
    with open(SCHEDULES_FILE, 'w') as f:
        json.dump({}, f)

if not os.path.exists(USERS_FILE):
    with open(USERS_FILE, 'w') as f:
        # Create a default admin user
        default_admin = {
            "username": "admin",
            "password_hash": hashlib.sha256("admin123".encode()).hexdigest(),
            "is_admin": True
        }
        json.dump([default_admin], f)

# Authentication functions
def load_users():
    try:
        with open(USERS_FILE, 'r') as f:
            return json.load(f)
    except:
        return []

def save_users(users):
    with open(USERS_FILE, 'w') as f:
        json.dump(users, f)

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password, password_hash):
    return hash_password(password) == password_hash

def get_user(username):
    users = load_users()
    for user in users:
        if user["username"] == username:
            return user
    return None

def authenticate_user(username, password):
    user = get_user(username)
    if not user:
        return False
    if not verify_password(password, user["password_hash"]):
        return False
    return user

# Helper functions
def load_engineers():
    try:
        with open(ENGINEERS_FILE, 'r', encoding='utf-8') as f:
            engineers = json.load(f)
            print(f"LOAD_ENGINEERS: Successfully loaded {len(engineers)} engineers from file")
            return engineers
    except Exception as e:
        print(f"LOAD_ENGINEERS ERROR: {str(e)}")
        return []

def save_engineers(engineers):
    try:
        if not isinstance(engineers, list):
            print(f"SAVE_ENGINEERS ERROR: Engineers is not a list, it's a {type(engineers)}")
            return

        if len(engineers) == 0:
            print("SAVE_ENGINEERS WARNING: Saving an empty engineers list!")
            
        # Create backup of current file if it exists
        if os.path.exists(ENGINEERS_FILE):
            backup_file = f"{ENGINEERS_FILE}.bak"
            shutil.copy2(ENGINEERS_FILE, backup_file)
            print(f"SAVE_ENGINEERS: Created backup at {backup_file}")
            
        print(f"SAVE_ENGINEERS: Saving {len(engineers)} engineers to file")
        # Print names of engineers being saved
        print(f"SAVE_ENGINEERS: Engineer names: {[eng.get('name', 'UNNAMED') for eng in engineers]}")
        
        with open(ENGINEERS_FILE, 'w', encoding='utf-8') as f:
            json.dump(engineers, f, indent=2, ensure_ascii=False)
            
        print(f"SAVE_ENGINEERS: Successfully saved {len(engineers)} engineers")
    except Exception as e:
        print(f"SAVE_ENGINEERS ERROR: {str(e)}")
        # Try to restore from backup if available
        backup_file = f"{ENGINEERS_FILE}.bak"
        if os.path.exists(backup_file):
            print(f"SAVE_ENGINEERS: Restoring from backup {backup_file}")
            shutil.copy2(backup_file, ENGINEERS_FILE)

def load_schedules():
    try:
        with open(SCHEDULES_FILE, 'r') as f:
            return json.load(f)
    except:
        return {}

def save_schedules(schedules):
    with open(SCHEDULES_FILE, 'w') as f:
        json.dump(schedules, f)

# Login required decorator
def login_required(f):
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

# Admin required decorator
def admin_required(f):
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('login_page'))
        if not session['user'].get('is_admin', False):
            flash('You need admin privileges to access this page.')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

# Routes
@app.route('/')
@login_required
def index():
    return render_template("index.html", 
                           workplaces=WORKPLACES, 
                           shifts=SHIFTS, 
                           username=session['user']['username'])

@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = authenticate_user(username, password)
        if not user:
            return render_template("login.html", error="Invalid username or password")
        
        # Store user in session
        session['user'] = {"username": user["username"], "is_admin": user["is_admin"]}
        
        return redirect(url_for('index'))
    
    # If GET or already logged in
    if 'user' in session:
        return redirect(url_for('index'))
    
    return render_template("login.html", error=None)

@app.route('/logout')
def logout():
    # Clear session
    session.clear()
    return redirect(url_for('login_page'))

@app.route('/admin')
@admin_required
def admin_page():
    # Remove password hashes for security
    users = load_users()
    users_display = []
    for user in users:
        users_display.append({
            "username": user["username"],
            "is_admin": user["is_admin"]
        })
    
    return render_template("admin.html", 
                           users=users_display, 
                           username=session['user']['username'])

@app.route('/admin/users', methods=['POST'])
@admin_required
def add_user():
    username = request.form.get('username')
    password = request.form.get('password')
    is_admin = True if request.form.get('is_admin') else False
    
    users = load_users()
    
    # Check if user already exists
    for user in users:
        if user["username"] == username:
            return render_template("admin.html", 
                                  error=f"User {username} already exists",
                                  users=[{"username": u["username"], "is_admin": u["is_admin"]} for u in users],
                                  username=session['user']['username'])
    
    # Create new user
    password_hash = hash_password(password)
    new_user = {
        "username": username,
        "password_hash": password_hash,
        "is_admin": is_admin
    }
    users.append(new_user)
    save_users(users)
    
    return redirect(url_for('admin_page'))

@app.route('/admin/users/<username>', methods=['DELETE'])
@admin_required
def delete_user(username):
    # Cannot delete yourself
    if username == session['user']['username']:
        return jsonify({"error": "Cannot delete yourself"}), 400
    
    users = load_users()
    users = [u for u in users if u["username"] != username]
    save_users(users)
    
    return jsonify({"status": "success"})

@app.route('/excel')
@login_required
def excel_version():
    """
    Alternative version of the main page with improved Excel generation capabilities.
    Use this URL if the main page is having issues.
    """
    return render_template("index.html", 
                          workplaces=WORKPLACES, 
                          shifts=SHIFTS, 
                          username=session['user']['username'])

@app.route('/excel-generator')
@login_required
def excel_generator_page():
    """
    Standalone page for generating Excel files from the schedule.
    Use this if the main Excel generation functionality isn't working.
    """
    # Get current Jalali date
    now_gregorian = datetime.now()
    now_jalali = jdt.datetime.fromgregorian(datetime=now_gregorian)
    current_jalali_year = now_jalali.year
    current_jalali_month = now_jalali.month
    
    return render_template("excel_generator.html", 
                          username=session['user']['username'],
                          current_jalali_year=current_jalali_year,
                          current_jalali_month=current_jalali_month,
                          persian_month_names=PERSIAN_MONTH_NAMES # Pass month names
                          )

# API Routes
@app.route('/api/engineers', methods=['GET'])
@login_required
def get_engineers():
    return jsonify(load_engineers())

@app.route('/api/engineers', methods=['POST'])
@admin_required
def add_engineer():
    data = request.json
    print(f"ADD_ENGINEER: Received data for engineer: {data['name']}")
    
    # Make a local copy of all engineers to avoid reference issues
    engineers = load_engineers()
    print(f"ADD_ENGINEER: Loaded {len(engineers)} existing engineers")
    
    # Check if updating or adding new
    engineer_exists = False
    engineer_index = -1
    
    # First, find if the engineer exists and get its index
    for i, eng in enumerate(engineers):
        if eng['name'] == data['name']:
            engineer_exists = True
            engineer_index = i
            break
    
    if engineer_exists:
        print(f"ADD_ENGINEER: Updating existing engineer at index {engineer_index}: {data['name']}")
        # Create a new dict for the updated engineer
        updated_engineer = {
            'name': data['name'],
            'workplaces': data['workplaces'],
            'limitations': data.get('limitations', {}),
            'minShifts': data.get('minShifts', 10),
            'maxShifts': data.get('maxShifts', 30)
        }
        # Replace the old engineer with the updated one
        engineers[engineer_index] = updated_engineer
    else:
        print(f"ADD_ENGINEER: Adding new engineer: {data['name']}")
        # Add the new engineer
        new_engineer = {
            'name': data['name'],
            'workplaces': data['workplaces'],
            'limitations': data.get('limitations', {}),
            'minShifts': data.get('minShifts', 10),
            'maxShifts': data.get('maxShifts', 30)
        }
        engineers.append(new_engineer)
    
    print(f"ADD_ENGINEER: Final list contains {len(engineers)} engineers")
    print(f"ADD_ENGINEER: Engineer names: {[eng.get('name', 'UNNAMED') for eng in engineers]}")
    
    # Make sure we're saving a copy to avoid any reference issues
    save_engineers(engineers[:])
    
    # Verify engineers were saved correctly
    verification = load_engineers()
    print(f"ADD_ENGINEER: Verification loaded {len(verification)} engineers")
    print(f"ADD_ENGINEER: Verified engineer names: {[eng.get('name', 'UNNAMED') for eng in verification]}")
    
    return jsonify({"status": "success"})

@app.route('/api/engineers/<n>', methods=['DELETE'])
@admin_required
def delete_engineer(n):
    engineers = load_engineers()
    engineers = [eng for eng in engineers if eng['name'] != n]
    save_engineers(engineers)
    return jsonify({"status": "success"})

@app.route('/api/schedule', methods=['GET'])
@login_required
def get_schedule():
    # Default to current Jalali year and month
    now_gregorian = datetime.now()
    now_jalali = jdt.datetime.fromgregorian(datetime=now_gregorian)
    
    year = request.args.get('year', default=now_jalali.year, type=int)
    month = request.args.get('month', default=now_jalali.month, type=int)
    
    schedules = load_schedules()
    # Use Jalali year/month for the key
    period_key = f"{year}-{month}" 
    
    if period_key in schedules:
        return jsonify(schedules[period_key])
    return jsonify({})

@app.route('/api/schedule', methods=['POST'])
@admin_required
def save_schedule():
    data = request.json
    # Assume year/month received from frontend are Jalali
    year = data.get('year')
    month = data.get('month')
    # Get the complete workplaces data from the request
    workplaces_data_from_request = data.get('workplaces', {})

    schedules = load_schedules()

    # Use Jalali year/month for the key
    period_key = f"{year}-{month}" 

    # *** FIX: Directly replace the data for the period ***
    # Instead of iterating and merging, just assign the received data.
    # This ensures that if the frontend sends a complete (potentially empty)
    # structure for the month, it fully overwrites whatever was there before.
    schedules[period_key] = workplaces_data_from_request

    save_schedules(schedules)
    return jsonify({"status": "success"})

@app.route('/api/generate_excel', methods=['POST'])
@login_required
def generate_excel():
    data = request.json
    # Assume year/month received from frontend are Jalali
    year = data.get('year')
    month = data.get('month')
    
    schedules = load_schedules()
    # Use Jalali year/month for the key
    period_key = f"{year}-{month}"
    
    if period_key not in schedules:
        return jsonify({"error": "No schedule data found for selected period"}), 404
    
    # Generate Excel files for each workplace
    excel_files = []
    for workplace in WORKPLACES:
        # Use Jalali year/month in filename
        filename = f"{workplace.replace(' ', '_')}_{year}_{month}.xlsx"
        file_path = os.path.join(DATA_DIR, filename)
        
        # Create Excel with formatting using Jalali calendar info
        create_excel_schedule(file_path, workplace, year, month, schedules[period_key].get(workplace, {}))
        excel_files.append(filename)
    
    return jsonify({
        "status": "success",
        "files": excel_files
    })

@app.route('/api/download/<filename>')
@login_required
def download_file(filename):
    file_path = os.path.join(DATA_DIR, filename)
    if not os.path.exists(file_path):
        return jsonify({"error": "File not found"}), 404
    return send_file(file_path, as_attachment=True)

@app.route('/api/pattern/upload', methods=['POST'])
@admin_required
def upload_pattern():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400
        
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "Invalid file format. Only Excel files (.xlsx, .xls) are supported."}), 400
    
    # Create a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        # Save the uploaded file
        file.save(tmp.name)
        tmp_path = tmp.name
    
    try:
        # Open the Excel file
        workbook = openpyxl.load_workbook(tmp_path, data_only=True)
        
        # Assume the first sheet is the pattern sheet
        sheet = workbook.active
        
        # Parse the pattern (expecting days as rows and shifts as columns)
        pattern = {}
        
        # Determine max rows to read (30 or 31 days)
        max_rows = min(sheet.max_row, 31)
        
        # Read each day (row) and the 3 shifts (columns)
        for day in range(1, max_rows + 1):
            pattern[str(day)] = {}
            
            # Read up to 3 shifts (columns)
            for shift in range(1, min(4, sheet.max_column + 1)):
                cell_value = sheet.cell(row=day, column=shift).value
                
                # Only include non-empty cells
                if cell_value:
                    pattern[str(day)][f"shift{shift}"] = str(cell_value).strip()
        
        return jsonify({
            "status": "success", 
            "pattern": pattern
        })
    except Exception as e:
        return jsonify({"error": f"Error processing Excel file: {str(e)}"}), 500
    finally:
        # Clean up the temporary file
        os.remove(tmp_path)

def create_excel_schedule(file_path, workplace, year, month, schedule_data):
    try:
        # Ensure year and month are integers
        year_int = int(year)
        month_int = int(month)

        # --- Calculate num_days manually to bypass potential .daysinmonth bug ---
        if 1 <= month_int <= 6:
            num_days = 31
        elif 7 <= month_int <= 11:
            num_days = 30
        elif month_int == 12:
            if jdt.isleap(year_int):
                num_days = 30
            else:
                num_days = 29
        else:
            raise ValueError(f"Invalid month number: {month_int}")
        # --- End manual calculation ---
        
        # Get month name 
        jalali_month_name = PERSIAN_MONTH_NAMES[month_int]

    except ValueError as ve:
        print(f"Error: Invalid integer year/month or invalid Jalali date: {year}-{month}. Details: {ve}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = f"ValueError for {year}-{month}: {ve}"
        wb.save(file_path)
        return
    # Removed other specific exception handlers (TypeError, AttributeError, generic Exception) related to the initial j_date_start
    # The code will now rely on Flask's default error handling if something unexpected happens after the ValueError check.

    # --- If we reach here, num_days and jalali_month_name should be set correctly ---

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{workplace} Schedule"
    
    # Define styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    weekend_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(bold=True, color="FFFFFF")
    centered = Alignment(horizontal='center', vertical='center')
    
    # Create title using Jalali month name and year
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    # Use the jalali_month_name obtained safely from the try block
    title_cell.value = f"{workplace} - {jalali_month_name} {year_int}" 
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = centered
    
    # Create headers
    headers = ["Day", "Shift 1", "Shift 2", "Shift 3"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = centered
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Fill in days (using num_days obtained safely from the try block)
    for day in range(1, num_days + 1):
        row = day + 3
        
        # Get Jalali date and Gregorian date for weekday calculation
        try:
            # Use integer year/month
            j_date = jdt.date(year_int, month_int, day) 
            
            g_date = j_date.togregorian()
            
            # Get Persian day name 
            gregorian_weekday = g_date.weekday()
            
            persian_weekday_index = (gregorian_weekday + 1) % 7 
            persian_day_name = PERSIAN_DAY_NAMES[persian_weekday_index]
            
        except ValueError as loop_ve:
            print(f"Error in day loop (ValueError): Could not process date {year_int}-{month_int}-{day}. Error: {loop_ve}")
            persian_day_name = "خطا"
            g_date = None 
        except Exception as loop_e: # Catch any unexpected error in the loop
            print(f"Error in day loop (Exception): Failed on day {day}. Error: {loop_e}, Type: {type(loop_e)}")
            # Optionally re-raise or handle more gracefully 
            # For now, just print and continue with error values
            persian_day_name = "خطا"
            g_date = None 
        
        # Day column with Persian day name
        day_cell = ws.cell(row=row, column=1)
        day_cell.value = f"{day} - {persian_day_name}" 
        day_cell.border = border
        day_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Weekend formatting
        is_weekend = (g_date.weekday() >= 5) if g_date else False 
        if is_weekend:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = weekend_fill
        
        # Fill in shifts
        day_str = str(day)
        if day_str in schedule_data:
            for shift_idx, shift in enumerate(SHIFTS, 1):
                shift_key = f"shift{shift_idx}"
                if shift_key in schedule_data[day_str]:
                    cell = ws.cell(row=row, column=shift_idx + 1)
                    cell.value = schedule_data[day_str][shift_key]
                    cell.border = border
                    cell.alignment = centered
                else:
                    ws.cell(row=row, column=shift_idx + 1).border = border
        else:
            for shift_idx in range(1, 4):
                ws.cell(row=row, column=shift_idx + 1).border = border
    
    # Set row height
    for row in range(1, num_days + 5):
        ws.row_dimensions[row].height = 25
    
    # Save workbook
    wb.save(file_path)

if __name__ == '__main__':
    app.run(debug=True, port=8000)